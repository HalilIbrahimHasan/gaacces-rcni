from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from azure_reconciliation import business_ready_sisense_reconciliation as recon


def _raw_rows() -> pd.DataFrame:
    base = {column: None for column in recon.RAW_COLUMNS}
    rows = []
    for values in (
        {
            "id": 1,
            "row_number_in_file": 1,
            "policy_id": "P1",
            "member_id": "M1",
            "subscriber_id": "M1",
            "enrolleeStatus": "CONFIRM",
        },
        {
            "id": 2,
            "row_number_in_file": 2,
            "policy_id": "P1",
            "member_id": "M1",
            "subscriber_id": "M1",
            "enrolleeStatus": "CONFIRM",
        },
        {
            "id": 3,
            "row_number_in_file": 3,
            "policy_id": "P2",
            "member_id": "M2",
            "subscriber_id": "S2",
            "enrolleeStatus": "TERM",
        },
        {
            "id": 4,
            "row_number_in_file": 4,
            "health_coverage_policy_no": "P3",
            "issuer_indiv_identifier": "M3",
            "subscriber_id": "S3",
            "enrolleeStatus": "CONFIRM",
        },
    ):
        row = {
            **base,
            "issuer": "99999",
            "year": 2026,
            "month": 1,
            "folder_year": 2026,
            "folder_month": 1,
            "filename_file_year": 2026,
            "filename_file_month": 1,
            "source_file": "sample_2026-01-05.xml",
            "source_file_path": "/tmp/sample_2026-01-05.xml",
            "file_hash": "hash",
            "loaded_at": pd.Timestamp("2026-01-05"),
            "coverage_year": 2026,
            "insurance_type": "Health",
            "insurance_type_code": "Health",
            "benefit_effective_date": pd.Timestamp("2026-01-01"),
            "member_maint_effective_date": pd.Timestamp("2026-01-05"),
            "action_code": "",
            "maintenance_type_code": "",
            **values,
        }
        rows.append(row)
    return pd.DataFrame(rows)


def test_sisense_reference_arithmetic_and_unique_dimensions() -> None:
    reference = recon.sisense_reference()

    assert len(reference) == 37
    assert not reference.duplicated(recon.DIMENSIONS).any()
    assert (
        reference["sisense_enrollment_total"]
        == reference["sisense_effectuated_enrollments"]
        + reference["sisense_pending_enrollments"]
    ).all()
    assert (
        reference["sisense_enrollee_total"]
        == reference["sisense_effectuated_enrollees"]
        + reference["sisense_pending_enrollees"]
    ).all()


def test_azure_statements_are_select_only() -> None:
    statements = [
        recon.HIGH_WATER_SQL,
        recon.COUNT_SQL,
        recon.ISSUER_YEAR_SQL,
        recon.RAW_EXTRACT_SQL,
    ]
    forbidden = (" insert ", " update ", " delete ", " merge ", " alter ", " drop ", " truncate ")
    for statement in statements:
        normalized = f" {' '.join(str(statement).lower().split())} "
        assert normalized.lstrip().startswith("select ")
        assert not any(keyword in normalized for keyword in forbidden)


def test_stage_pipeline_uses_production_fallback_identifiers() -> None:
    snapshots, raw_confirm, collapse_audit, annual_distinct = recon.process_issuer_year(
        _raw_rows(),
        issuer="99999",
        reporting_year=2026,
    )
    combined = pd.concat(snapshots, ignore_index=True)
    raw = combined[combined["stage_name"] == "01_raw_annual_distinct"].iloc[0]

    assert list(dict.fromkeys(combined["stage_name"])) == recon.STAGE_ORDER
    assert int(raw["output_row_count"]) == 4
    assert int(raw["enrollment_count"]) == 3
    assert int(raw["enrollee_count"]) == 3
    assert int(raw_confirm.iloc[0]["raw_confirm_enrollments"]) == 2
    assert int(raw_confirm.iloc[0]["raw_confirm_enrollees"]) == 2
    assert len(collapse_audit) == 1
    assert int(annual_distinct.iloc[0]["enrollment_count"]) == 3


def test_run_reconciliation_writes_required_sheets(
    monkeypatch,
    tmp_path: Path,
) -> None:
    raw = _raw_rows()
    monkeypatch.setattr(recon, "fetch_high_water_id", lambda engine: 4)
    monkeypatch.setattr(
        recon,
        "validate_source_counts",
        lambda engine, years, expected, high_water_id: (
            pd.DataFrame([{"folder_year": 2026, "row_count": len(raw)}]),
            [{
                "check_name": "source_row_count_2026",
                "expected": len(raw),
                "actual": len(raw),
                "passed": True,
                "severity": "ERROR",
            }],
        ),
    )
    monkeypatch.setattr(
        recon,
        "fetch_issuer_years",
        lambda engine, years, high_water_id: pd.DataFrame([{
            "issuer": "99999",
            "folder_year": 2026,
            "row_count": len(raw),
        }]),
    )
    output = tmp_path / "business_ready_sisense_reconciliation.xlsx"
    result = recon.run_reconciliation(
        object(),
        years=[2026],
        output_path=output,
        expected_counts={2026: len(raw)},
        raw_fetcher=lambda engine, issuer, year, high_water_id: raw,
    )

    assert output.exists()
    required = {
        "run_summary",
        "stage_waterfall",
        "stage_by_issuer",
        "final_comparison",
        "ours_only",
        "sisense_only",
        "largest_gaps",
        "raw_confirm_vs_effectuated",
        "unresolved_definitions",
        "data_quality_checks",
    }
    assert required.issubset(load_workbook(output, read_only=True).sheetnames)
    assert set(result.stage_by_issuer["stage_name"]) == set(recon.STAGE_ORDER)
    assert (
        result.raw_confirm_vs_effectuated["reconciliation_status"]
        == "UNRESOLVED_VIMO_PREDICATE"
    ).all()
